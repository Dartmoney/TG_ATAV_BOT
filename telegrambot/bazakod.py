import openpyxl
from fuzzywuzzy import fuzz

# Открываем Excel-файл
workbook = openpyxl.load_workbook('example.xlsx')

# Выбираем лист с данными
sheet = workbook['Лист1']  # Замените 'Лист1' на имя вашего листа

# dicti = {}
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     dicti[row[0]] = dicti[row[1]]

dicti = {}


def find(name):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if fuzz.WRatio(row[0], name) > 70:
            dicti[row[0]] = row[1]
    if dicti:
        return dicti
    return "вопрос не найден"

# Пример использования
